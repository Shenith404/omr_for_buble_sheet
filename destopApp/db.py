import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
class OMRJsonHandler:
    def __init__(self, project_path):
        self.json_path = os.path.join(project_path, "results", "answers.json")

        os.makedirs(os.path.dirname(self.json_path), exist_ok=True)
        self._initialize_file()

    def _initialize_file(self):
        """Create empty JSON file if not exists"""
        if not os.path.exists(self.json_path):
            with open(self.json_path, 'w') as f:
                json.dump({
                    "_metadata": {
                        "version": 1.0,
                        "created": datetime.now().isoformat(),
                        "total_sheets": 0
                    },
                    "data": {}
                }, f, indent=4)

    def _load_data(self):
        """Load and validate JSON data"""
        with open(self.json_path, 'r') as f:
            data = json.load(f)
            
            # Backward compatibility for old format
            if "data" not in data:
                data = {"_metadata": {}, "data": data}
                
            return data

    def _save_data(self, data):
        """Atomic write to JSON file"""
        temp_path = self.json_path + ".tmp"
        with open(temp_path, 'w') as f:
            json.dump(data, f, indent=4)
        os.replace(temp_path, self.json_path)

    # --- CRUD Operations ---
    def create_or_update_sheet(self, filename, detected_answers, total_marks=0):
        data = self._load_data()
        #detected answers saved as  1->2 2->3 3->4 4->5 and no answers saved as -1


        data["data"][filename] = {
            "detected": detected_answers,
            "total_marks": total_marks,
            "reviewed": False,
            "last_modified": datetime.now().isoformat()
        }

        data["_metadata"]["total_sheets"] = len(data["data"])
        data["_metadata"]["last_updated"] = datetime.now().isoformat()

        self._save_data(data)


    def get_sheet(self, filename):
        """Read single sheet data"""
        data = self._load_data()
        return data["data"].get(filename)

    def get_all_sheets(self, reviewed=None):
        """Get all sheets with optional filter"""
        data = self._load_data()
        if reviewed is None:
            return data["data"]
        
        return {
            k: v for k, v in data["data"].items()
            if v["reviewed"] == reviewed
        }

    def update_correction(self, filename, question_num, new_answer):
        """Manually correct a single answer"""
        data = self._load_data()
        
        if filename not in data["data"]:
            raise ValueError("Sheet not found")
            
        sheet = data["data"][filename]
        
        # Initialize 'detected' if it doesn't exist
        if "detected" not in sheet:
            sheet["detected"] = {}
        
        sheet["detected"][question_num] = new_answer
        sheet["reviewed"] = False
        sheet["last_modified"] = datetime.now().isoformat()
        
        # Recalculate marks if needed
        if "model_answers" in data["_metadata"]:
            sheet["total_marks"] = self._calculate_marks(
                sheet["detected"],
                data["_metadata"]["model_answers"]
            )
        
        self._save_data(data)
        return sheet["detected"]
    
    

    def delete_sheet(self, filename):
        """Remove a sheet record"""
        data = self._load_data()
        if filename in data["data"]:
            del data["data"][filename]
            data["_metadata"]["total_sheets"] = len(data["data"])
            self._save_data(data)
            return True
        return False
        
    def delete_answers_file(self):
        """Delete and recreate the answers.json file"""
        try:
            if os.path.exists(self.json_path):
                os.remove(self.json_path)
            self._initialize_file()  # Recreate the file
            return True
        except Exception as e:
            print(f"Error deleting and recreating answers file: {e}")
            return False

    # --- Helper Methods ---
    def set_model_answers(self, answers):
        """Store model answers for automatic marking"""
        data = self._load_data()
        data["_metadata"]["model_answers"] = answers
        self._save_data(data)

    def save_model_answers(self, answers):
        """Save model answers to the JSON file"""
        data = self._load_data()
        data["_metadata"]["model_answers"] = answers
        self._save_data(data)

    def read_model_answers(self):
        """Read model answers from the JSON file"""
        data = self._load_data()
        return data["_metadata"].get("model_answers")

    def _calculate_marks(self, student_answers, model_answers):
        """Calculate total correct answers"""
        return sum(1 for s, m in zip(student_answers, model_answers) if s-1 == m)

    def mark_for_review(self, filename, status=True):
        """Flag/unflag a sheet for manual review"""
        data = self._load_data()
        if filename in data["data"]:
            data["data"][filename]["reviewed"] = status
            self._save_data(data)
            return True
        return False
    
    #get reviewed sheet names
    def get_reviewed_sheet_names(self):
        """Get names of all reviewed sheets"""
        data = self._load_data()
        return [k for k, v in data["data"].items() if v.get("reviewed", True)]
    

    #save answers in excel sheet
    def export_to_excel(self, save_path="omr_export.xlsx"):
        """Export all sheet data to an Excel file with formatting"""
        data = self._load_data()
        all_sheets = data.get("data", {})
        model_answers = data.get("_metadata", {}).get("model_answers", [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # --- Define styles ---
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue
        reviewed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        unreviewed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        center_alignment = Alignment(horizontal="center")

        # --- Header Row ---
        headers = ["Filename", "Reviewed", "Total Marks", "Last Modified", "Detected Answers"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = center_alignment

        # --- Fill Rows ---
        for row_num, (filename, details) in enumerate(all_sheets.items(), start=2):
            ws.cell(row=row_num, column=1, value=filename)
            ws.cell(row=row_num, column=2, value=str(details.get("reviewed", False)))
            ws.cell(row=row_num, column=3, value=details.get("total_marks", 0))
            ws.cell(row=row_num, column=4, value=details.get("last_modified", ""))
            detected = details.get("detected", {})
            ws.cell(row=row_num, column=5, value=str(detected))

            # Apply conditional row coloring
            fill = reviewed_fill if details.get("reviewed") else unreviewed_fill
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = center_alignment

        # --- Auto-adjust column widths ---
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        # --- Save workbook ---
        try:
            wb.save(save_path)
            print(f"Excel file exported successfully to '{save_path}'")
            return True
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
