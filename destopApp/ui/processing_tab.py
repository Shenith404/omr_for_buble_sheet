import json
import os
import csv
import cv2
import numpy as np
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QPushButton, QLabel, QFileDialog, QProgressBar,
    QMessageBox, QSizePolicy, QApplication  # Added QApplication
)
from PySide6.QtGui import QPixmap, QImage
from PySide6.QtCore import Qt, Signal, QThread, QObject

import utils
import model
import subprocess  # For opening the CSV file directly
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation
import db


class OMRProcessor(QObject):
    """
    Optimized OMR processing worker that handles batch processing of answer sheets
    Features:
    - Thread-safe operation
    - Progress reporting
    - Error handling
    - Memory-efficient processing
    """
    
    # Signals for communication with main thread
    progress_updated = Signal(int, str)  # progress_percent, status_message
    image_processed = Signal(str, list, np.ndarray)  # filename, answers, marked_image
    processing_complete = Signal()
    error_occurred = Signal(str)
    cancelled = Signal()

    def __init__(self, image_paths, project_path, model_answers):
        super().__init__()
        self.image_paths = image_paths
        self.project_path = project_path
        self.model_answers = model_answers  # Model answers for comparison
        self._cancel_requested = False
        self.widthImg = 1025  # Standard OMR sheet width
        self.heightImg = 760  # Standard OMR sheet height
        self.batch_size = 50  # Process images in batches to manage memory
        self.dummy_answer = [0] * 50  # Placeholder for dummy answers
        self.total_marks = 0  # Initialize total marks

    def process_all(self):
        """Process all images with accurate progress tracking"""
        try:
            total_images = len(self.image_paths)
            processed_count = 0
            
            for image_path in self.image_paths:
                if self._cancel_requested:
                    self.cancelled.emit()
                    break
                
                filename = os.path.basename(image_path)
                
                try:
                    # Load image
                    image = cv2.imread(image_path)
                    if image is None:
                        raise ValueError(f"Failed to load image: {filename}")
                    
                    # Process image
                    detected_answers, final_img = self.process_omr_sheet(image)
                    
                    # Emit progress after completion (not before)
                    processed_count += 1
                    progress = int((processed_count / total_images) * 100)
                    self.progress_updated.emit(
                        progress,
                        f"Processed {filename} ({processed_count}/{total_images})"
                    )
                    
                    # Emit results
                    self.image_processed.emit(filename, detected_answers, final_img)
                    
                except Exception as e:
                    self.error_occurred.emit(f"Skipped {filename}: {str(e)}")
                    continue
            
            if not self._cancel_requested:
                self.processing_complete.emit()
        
        except Exception as e:
            self.error_occurred.emit(f"Fatal processing error: {str(e)}")

    def process_omr_sheet(self, image):
        """
        Process a single OMR sheet with optimized operations
        Returns:
        - answers: List of detected answers (1-based index)
        - marked_image: Image with marked answers
        """
        # Step 1: Preprocessing with optimized operations
        img = cv2.resize(image, (self.widthImg, self.heightImg))
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 1)
        edges = cv2.Canny(blur, 10, 50)

        #disable  mark button when processing
        


     
        if self._cancel_requested:
            raise RuntimeError("Processing cancelled")

        # Step 2: Contour Detection with area filtering
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        rects = utils.rectContour(contours)
        
        if not rects:
            return self.dummy_answer, img

        # Step 3: Perspective Transform with error checking
        biggest = utils.getCornerPoints(rects[0])
        if biggest.size == 0:
            return self.dummy_answer, img
            
        biggest = utils.reorder(biggest)
        pts1 = np.float32(biggest)
        pts2 = np.float32([[0, 0], [self.widthImg, 0], [0, self.heightImg], [self.widthImg, self.heightImg]])
        matrix = cv2.getPerspectiveTransform(pts1, pts2)
        warped = cv2.warpPerspective(img, matrix, (self.widthImg, self.heightImg))

        # Step 4: Adaptive Thresholding for better robustness
        warped_gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
        #increase brightness
        #warped_gray =cv2.convertScaleAbs(warped_gray, alpha=1, beta=50)
        thresh = cv2.threshold(warped_gray, 170, 255, cv2.THRESH_BINARY_INV)[1]

       # totalPixelSize =cv2.countNonZero(thresh)
       # print("Total Pixel Size: ",totalPixelSize)

        if self._cancel_requested:
            raise RuntimeError("Processing cancelled")

        # Step 1: Split into boxes
        boxes = utils.verticalSplitBoxes(warped_gray)
        thresh_boxes =utils.verticalSplitBoxes(thresh)
        detected_answers = []

        # Step 2: For each question box
        for i,box in enumerate(boxes):
            if self._cancel_requested:
                raise RuntimeError("Processing cancelled")

            # Get answer bubbles (skip unwanted blocks)
            answer_blocks = utils.getAnswerBlocks(boxes[i])[2:6]  # Adjust if needed
            thresh_answer_blocks =utils.getAnswerBlocks(thresh_boxes[i])[2:6]
            
            #if total pixel value of thresh_answer_block is greater than 240 replace relevent anwer_block by increasing constrast
            for j, tab in enumerate(thresh_answer_blocks):
                p_val = cv2.countNonZero(tab)
                if p_val > 1700:  # Careful: You had a typo (944480), it should be 255*64*64 = 1044480
                    print("psfds", p_val)
                    answer_blocks[j] = cv2.convertScaleAbs(answer_blocks[j], alpha=1.5, beta=50)

            # Step 3: Collect non-empty blocks for batch processing
            valid_blocks = [(j, block) for j, block in enumerate(answer_blocks) if cv2.countNonZero(block) > 0]

            if valid_blocks:
                indices, blocks = zip(*valid_blocks)  # unzip indices and images

                # Step 4: Batch classify
                predictions = model.classify_batch(list(blocks))  # returns list of (label, confidence)

                # Step 5: Get crossed bubble index (if only one)
                marked = []
                for idx, (label, _) in zip(indices, predictions):
                    if label == "cross_Images":
                        marked.append(idx + 2)  # 1-based index since we skipped first 2 blocks

                detected_answers.append(marked[0] if len(marked) == 1 else -1)
            else:
                detected_answers.append(-1)  # No valid bubbles

        # Step 6: Generate Results with optimized drawing
        drawing = np.zeros_like(warped)
        drawing,self.total_marks = utils.showAnswers(drawing, detected_answers, self.model_answers)
        
        inv_matrix = cv2.getPerspectiveTransform(pts2, pts1)
        inv_drawing = cv2.warpPerspective(drawing, inv_matrix, (img.shape[1], img.shape[0]))
        final_img = cv2.addWeighted(img, 1, inv_drawing, 1, 0)
        cv2.putText(final_img, f"Total Marks: {self.total_marks}/50", (50, 700), cv2.FONT_HERSHEY_COMPLEX_SMALL, 2, (0, 0, 250), 1)


        return detected_answers, final_img

    def cancel(self):
        """Request cancellation of current processing"""
        self._cancel_requested = True


class ProcessingTab(QWidget):
    """
    Optimized Processing Tab for handling large batches of OMR sheets
    Features:
    - Memory-efficient image handling
    - Batch processing
    - Progress tracking
    - Error recovery
    """
    
    # Processing signals
    processing_complete = Signal(object, object, object)
    processing_cancelled = Signal()
    processing_started = Signal()
    processing_finished = Signal()
    
    def __init__(self):
        super().__init__()
        self.project_path = None
        self.image_paths = []
        self.current_index = 0
        self.processed_images = {}  # Stores only paths to processed images
        self.current_answers = {}  # Stores answers for CSV output
        self.processed_count = 0  # Track the number of processed images
        self.setup_ui()
        self.setup_connections()
        self.model_answers = []  # Placeholder for model answers
        self.handler=None
        
        # Initialize UI state
        self.image_label.setAlignment(Qt.AlignCenter)
        self.lbl_image_info.setText("0/0 images loaded")
        self.update_navigation_buttons()

    def setup_ui(self):
        """Initialize all UI components with optimized layouts"""
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(10, 10, 10, 10)

        # Project Selection Group
        project_group = QGroupBox("Project Selection")
        project_layout = QHBoxLayout()
        self.btn_select_project = QPushButton("Select Project")
        self.btn_select_project.setStyleSheet("font-weight: bold;")
        self.lbl_project = QLabel("No project selected")
        project_layout.addWidget(self.btn_select_project)
        project_layout.addWidget(self.lbl_project)
        project_group.setLayout(project_layout)

        # Image Navigation Group
        nav_group = QGroupBox("Image Navigation")
        nav_layout = QHBoxLayout()
        self.btn_prev = QPushButton("◀ Previous")
        self.btn_next = QPushButton("Next ▶")
        self.btn_delete_image = QPushButton("Delete Image")  # Move delete button here
        self.btn_delete_image.setStyleSheet("""
            background-color: #f44336; 
            color: white;
            font-weight: bold;
        """)  # Add red background
        self.btn_delete_image.clicked.connect(self.delete_current_image)
        self.lbl_image_info = QLabel("0/0 images loaded")
        nav_layout.addWidget(self.btn_prev)
        nav_layout.addWidget(self.lbl_image_info)
        nav_layout.addWidget(self.btn_next)
        nav_layout.addWidget(self.btn_delete_image)  # Add delete button to navigation layout
        nav_group.setLayout(nav_layout)

        # Image Display with optimized settings
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setStyleSheet("""
            border: 1px solid gray; 
            min-height: 400px;
            background-color: #f0f0f0;
        """)
        self.image_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Processing Controls with batch options
        control_group = QGroupBox("Batch Processing")
        control_layout = QVBoxLayout()
        self.btn_process_all = QPushButton("Mark All Images")
        self.btn_process_all.setStyleSheet("""
            background-color: #4CAF50; 
            color: white;
            padding: 8px;
            font-weight: bold;
        """)
        self.btn_cancel = QPushButton("Cancel Processing")
        self.btn_cancel.setStyleSheet("""
            background-color: #f44336;
            color: white;
            padding: 8px;
        """)
        self.btn_cancel.setEnabled(False)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setTextVisible(True)
        
        self.lbl_status = QLabel("Ready")
        self.lbl_status.setStyleSheet("font-weight: bold;")
        
        # Add a button to load model answers
        self.btn_save_model_answers = QPushButton("Save Answers")
        self.btn_save_model_answers.setStyleSheet("font-weight: bold; " )
        self.btn_save_model_answers.setEnabled(False)  # Initially disabled

        # Add new button for editing answers
        self.btn_edit_answers = QPushButton("Edit Answers")
        self.btn_edit_answers.setStyleSheet("font-weight: bold;")
        self.btn_edit_answers.setEnabled(False)  # Initially disabled

        control_layout.addWidget(self.btn_save_model_answers)  # Add to the control group
        control_layout.addWidget(self.btn_edit_answers)  # Add the new button
        control_layout.addWidget(self.btn_process_all)
        control_layout.addWidget(self.btn_cancel)
        control_layout.addWidget(self.progress_bar)
        control_layout.addWidget(self.lbl_status)
        control_group.setLayout(control_layout)

        # Assemble main layout
        self.layout.addWidget(project_group)
        self.layout.addWidget(nav_group)
        self.layout.addWidget(self.image_label)
        self.layout.addWidget(control_group)
        self.setLayout(self.layout)

    def setup_connections(self):
        """Connect all signals and slots"""
        self.btn_select_project.clicked.connect(self.select_project)
        self.btn_prev.clicked.connect(self.show_previous_image)
        self.btn_next.clicked.connect(self.show_next_image)
        self.btn_process_all.clicked.connect(self.start_processing)
        self.btn_cancel.clicked.connect(self.cancel_processing)
        self.btn_save_model_answers.clicked.connect(self.save_model_answers)
        self.btn_edit_answers.clicked.connect(self.edit_answers)  # Connect the new button

    def select_project(self):
        """Let user select a project folder with validation"""
        project_path = QFileDialog.getExistingDirectory(
            self, 
            "Select Project Folder",
            "",
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks
        )
        if project_path:
            self.load_project(project_path)

    def load_project(self, project_path):
        """Load project with optimized file handling"""
        self.project_path = project_path
        self.lbl_project.setText(os.path.basename(project_path))
        self.image_paths = []
        self.processed_images.clear()
        self.current_answers.clear()
        
        # Load images with error handling
        try:
            # Load model answers if available
            self.save_model_answers()
            # disable save answers button
            self.btn_save_model_answers.setEnabled(False)  # Disable the button after loading
            # Load from original_images folder
            images_dir = os.path.join(project_path, "original_images")
            if os.path.exists(images_dir):
                self.image_paths.extend([
                    os.path.join(images_dir, f) 
                    for f in sorted(os.listdir(images_dir))
                    if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))
                ])
            
            # Load linked images from references file
            ref_file = os.path.join(project_path, "image_references.txt")
            if os.path.exists(ref_file):
                with open(ref_file, 'r') as f:
                    self.image_paths.extend([
                        line.strip() 
                        for line in f 
                        if line.strip() and os.path.exists(line.strip())
                    ])
        
        except Exception as e:
            QMessageBox.warning(self, "Load Error", f"Error loading project: {str(e)}")
        
        # Update UI
        if self.image_paths:
            self.current_index = 0
            self.show_current_image()
        else:
            self.lbl_image_info.setText("No valid images found")
        
        self.update_navigation_buttons()
        self.btn_save_model_answers.setEnabled(True)  # Enable the button when a project is loaded
        self.btn_edit_answers.setEnabled(True)  # Enable the edit button when a project is loaded

    def show_current_image(self):
        """Display the current image without showing marked answers"""
        if not self.image_paths:
            self.image_label.clear()
            self.lbl_image_info.setText("No images loaded")
            return

        try:
            image_path = self.image_paths[self.current_index]
            filename = os.path.basename(image_path)

            # Load the original image
            image = cv2.imread(image_path)
            if image is None:
                raise ValueError("Failed to read image")

            # Convert the image from BGR (OpenCV format) to RGB (Qt format)
            rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_image.shape
            bytes_per_line = ch * w

            # Create a QImage object from the RGB image data
            q_img = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(q_img)

            # Scale the pixmap to fit the QLabel while maintaining the aspect ratio
            scaled_pixmap = pixmap.scaled(
                self.image_label.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )

            # Set the scaled pixmap to the QLabel for display
            self.image_label.setPixmap(scaled_pixmap)
            self.lbl_image_info.setText(
                f"Image {self.current_index + 1}/{len(self.image_paths)}\n"
                f"{filename}"
            )

        except Exception as e:
            # Handle any errors that occur during image loading or display
            print(f"Image display error: {str(e)}")
            self.image_label.setText(f"Error: {str(e)}")
            self.lbl_image_info.setText("Image display error")

    def show_next_image(self):
        """Navigate to the next image with bounds checking"""
        if self.current_index < len(self.image_paths) - 1:
            self.current_index += 1
            self.show_current_image()
        self.update_navigation_buttons()

    def show_previous_image(self):
        """Navigate to the previous image with bounds checking"""
        if self.current_index > 0:
            self.current_index -= 1
            self.show_current_image()
        self.update_navigation_buttons()
            
    def update_navigation_buttons(self):
        """Update button states based on current position"""
        has_images = len(self.image_paths) > 0
        self.btn_prev.setEnabled(has_images and self.current_index > 0)
        self.btn_next.setEnabled(has_images and self.current_index < len(self.image_paths) - 1)
        self.btn_process_all.setEnabled(has_images)

    def start_processing(self):
        """Start optimized batch processing"""

        # give and exception when save answers button is enabled
        if self.btn_save_model_answers.isEnabled():
            QMessageBox.warning(self, "Error", "Please check & save model answers before processing.")
            return

        if not self.image_paths or not self.project_path:
            QMessageBox.warning(self, "Error", "No project or images loaded")
            return
        try:
            # Initialize results directory
            results_dir = os.path.join(self.project_path, "results")
            os.makedirs(results_dir, exist_ok=True)
            
            # Initialize CSV file
           # csv_path = os.path.join(results_dir, "answers.csv")

            # Initialize json file
            self.handler = db.OMRJsonHandler(self.project_path)
            #delete existing json file
            self.handler.delete_answers_file()

            #save model answers to json file
            self.handler.save_model_answers(self.model_answers)


            # with open(csv_path, 'w', newline='') as csvfile:
            #     writer = csv.writer(csvfile)
            #     writer.writerow(["Filename"] + [f"Q{i+1}" for i in range(50)])
            
            # Reset processing state
            self.processed_images.clear()
            self.current_answers.clear()
            self.processed_count = 0  # Reset processed count
            
            # Setup worker thread with lower priority
            self.processing_started.emit()
            self.btn_process_all.setEnabled(False)
            self.btn_cancel.setEnabled(True)
            self.btn_delete_image.setEnabled(False)  # Disable delete button
            self.progress_bar.setValue(0)
            self.lbl_status.setText("Initializing batch processing...")
            QApplication.processEvents()  # Ensure UI updates
            
            self.worker_thread = QThread()
            self.worker_thread.setPriority(QThread.LowPriority)
            self.omr_processor = OMRProcessor(self.image_paths, self.project_path, self.model_answers)
            self.omr_processor.moveToThread(self.worker_thread)
            
            # Connect signals
            self.worker_thread.started.connect(self.omr_processor.process_all)
            self.omr_processor.progress_updated.connect(self.update_progress)
            self.omr_processor.image_processed.connect(
                lambda filename, answers, marked_image: self.save_image_results(filename, answers, marked_image, self.omr_processor.total_marks)
            )
            self.omr_processor.processing_complete.connect(self.finish_processing)
            self.omr_processor.error_occurred.connect(self.handle_processing_error)
            self.omr_processor.cancelled.connect(self.cancel_processing)
            
            # Cleanup connections
            self.omr_processor.processing_complete.connect(self.worker_thread.quit)
            self.omr_processor.error_occurred.connect(self.worker_thread.quit)
            self.worker_thread.finished.connect(self.worker_thread.deleteLater)
            
            # Start processing
            self.worker_thread.start()
        except Exception as e:
            QMessageBox.critical(self, "Processing Error", f"Error starting processing: {str(e)}")
            return

    def update_progress(self, progress, message):
        """Throttled progress updates"""
        # Only update if progress increased or it's a completion message
        if progress > self.progress_bar.value() or progress == 100:
            self.progress_bar.setValue(progress)
            self.lbl_status.setText(message)
            
            # Process events at certain intervals (every 5% or completion)
            if progress % 5 == 0 or progress == 100:
                QApplication.processEvents()  # Ensure UI remains responsive


    def save_image_results(self, filename, answers, marked_image,total_marks):
        """Save results immediately after each image is processed"""
        if not self.project_path or marked_image is None:
            return
            
        results_dir = os.path.join(self.project_path, "results")
        
        try:
            # 1. Ensure directory exists
            os.makedirs(results_dir, exist_ok=True)
            
            # 2. Save marked image
            output_path = os.path.join(results_dir, f"{filename}")
            cv2.imwrite(output_path, marked_image)
            
            # 3. Save answers to CSV immediately
            # csv_path = os.path.join(results_dir, "answers.csv")
            # file_exists = os.path.exists(csv_path)
            
            # with open(csv_path, 'a', newline='') as csvfile:
            #     writer = csv.writer(csvfile)
            #     if not file_exists:
            #         writer.writerow(["Filename", "TotalMarks", ""] + [f"Q{i+1}" for i in range(len(answers))])
                
            #     # Convert answers to 1-based index (0 for unanswered)
            #     corrected_answers = [ans-1 if ans != -1 else 0 for ans in answers]
            #     writer.writerow([filename, total_marks, ""] + corrected_answers)
            
            
            # 4. Save answers in a json file
            #self.handler = db.OMRJsonHandler(self.project_path)

            self.handler.create_or_update_sheet(filename, answers, total_marks)
                    
            
            # 4. Update UI
            self.display_marked_image(marked_image)
            self.processed_count += 1  # Increment processed count
            if self.current_index < len(self.image_paths) - 1:
                self.current_index += 1
            self.update_navigation_buttons()

            
            
        except Exception as e:
            self.handle_processing_error(f"Error saving {filename}: {str(e)}")

    def display_marked_image(self, marked_image):
        """Display marked image with optimized rendering"""
        try:
            # Convert color space
            rgb_image = cv2.cvtColor(marked_image, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_image.shape
            bytes_per_line = ch * w
            
            # Create QImage
            q_img = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(q_img)
            
            # Scale to fit
            scaled_pixmap = pixmap.scaled(
                self.image_label.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            
            self.image_label.setPixmap(scaled_pixmap)
            self.lbl_image_info.setText(
                f"Processed Image {self.current_index + 1}/{len(self.image_paths)}\n"
                f"Marked results displayed"
            )
            
        except Exception as e:
            print(f"Marked image display error: {str(e)}")
            self.image_label.setText(f"Error displaying results")

    def finish_processing(self):
        """Verify all images were processed and allow navigation to other tabs"""
        total_count = len(self.image_paths)
        
        if self.processed_count < total_count:
            self.lbl_status.setText(
                f"Completed {self.processed_count}/{total_count} images"
            )
        else:
            self.lbl_status.setText("Processing completed successfully!")
        
        self.progress_bar.setValue(100)
        self.btn_process_all.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.btn_delete_image.setEnabled(True)  # Enable delete button

        # Ensure the worker thread is properly terminated
        if hasattr(self, 'worker_thread') and self.worker_thread.isRunning():
            self.worker_thread.quit()
            self.worker_thread.wait()

        QMessageBox.information(
            self,
            "Processing Complete",
            f"Processed {self.processed_count}/{total_count} images\n"
            f"Results saved to: {os.path.join(self.project_path, 'results')}"
        )

        # Emit a signal or update the UI to allow navigation
        self.processing_finished.emit()

    def cancel_processing(self):
        """Handle processing cancellation and reset state"""
        if hasattr(self, 'omr_processor'):
            self.omr_processor.cancel()

            # Ensure the worker thread is properly terminated
        if hasattr(self, 'worker_thread') and self.worker_thread.isRunning():
            self.worker_thread.quit()
            self.worker_thread.wait()


        # Reset processing state
        self.processed_images.clear()
        self.current_answers.clear()
        self.processed_count = 0  # Reset processed count
        self.progress_bar.setValue(0)
        self.btn_process_all.setEnabled(True)
        self.btn_cancel.setEnabled(False)

        # Emit a signal or update the UI to allow navigation
        self.processing_cancelled.emit()


    def handle_processing_error(self, error_msg):
        """Handle processing errors with user feedback"""
        self.btn_process_all.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.lbl_status.setText(f"Error: {error_msg}")
        self.progress_bar.setValue(0)
        
        # Show error message but don't block processing
        QMessageBox.critical(self, "Processing Error", error_msg)

    def resizeEvent(self, event):
        """Handle window resize to maintain image display"""
        super().resizeEvent(event)
        if hasattr(self, 'image_label') and self.image_paths:
            self.show_current_image()

    def set_image_paths(self, image_paths):
        """
        Set the image paths to process
        Args:
            image_paths: List of paths to images
        """
        # Validate input
        if not isinstance(image_paths, list):
            raise TypeError("image_paths must be a list")
        if not all(isinstance(p, (str, os.PathLike)) for p in image_paths):
            raise TypeError("All paths must be strings or PathLike objects")
        
        # Store paths and reset state
        self.image_paths = [os.path.normpath(str(p)) for p in image_paths]  # Normalize paths
        self.current_index = 0
        self.processed_images = {}
        
        # Update UI
        if self.image_paths:
            self.show_current_image()
        else:
            self.image_label.clear()
            self.lbl_image_info.setText("No images loaded")
        
        self.update_navigation_buttons()
        self.btn_process_all.setEnabled(len(self.image_paths) > 0)

    def closeEvent(self, event):
        """Clean up resources when closing"""
        if hasattr(self, 'worker_thread') and self.worker_thread.isRunning():
            self.cancel_processing()
            self.worker_thread.quit()
            self.worker_thread.wait(1000)  # Wait up to 1 second
        event.accept()
        self.btn_save_model_answers.setEnabled(False)  # Disable the button when the application is closed

    def showEvent(self, event):
        """Reload images every time the tab is shown"""
        super().showEvent(event)
        if self.project_path:
            self.load_project(self.project_path)

    def delete_current_image(self):
        """Delete the currently previewed image permanently from the project"""
        if not self.image_paths:
            QMessageBox.warning(self, "Delete Error", "No images to delete.")
            return

        try:
            # Get the current image path
            image_path = self.image_paths[self.current_index]
            filename = os.path.basename(image_path)

            # Confirm deletion
            reply = QMessageBox.question(
                self,
                "Delete Image",
                f"Are you sure you want to delete '{filename}' permanently?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

            # Delete the file from disk
            os.remove(image_path)

            # Remove the image from the list
            del self.image_paths[self.current_index]

            # Update the current index
            if self.current_index >= len(self.image_paths):
                self.current_index = max(0, len(self.image_paths) - 1)

            # Update the UI
            if self.image_paths:
                self.show_current_image()
            else:
                self.image_label.clear()
                self.lbl_image_info.setText("No images loaded")

            self.update_navigation_buttons()

            QMessageBox.information(self, "Delete Successful", f"'{filename}' has been deleted.")

        except Exception as e:
            QMessageBox.critical(self, "Delete Error", f"Failed to delete image: {str(e)}")

    def save_model_answers(self):
        """Load or create model answers XLSX file and open it for editing"""
        try:
            if not self.project_path:
                QMessageBox.warning(self, "Error", "No project is open. Please select a project first.")
                return

            # Define the path for the model answers XLSX file
            model_answers_path = os.path.join(self.project_path, "model_answers.xlsx")

            # Check if the file exists; if not, create it with default answers and a header
            if not os.path.exists(model_answers_path):
                self.create_model_answers_file(model_answers_path)

            # self.open_csv_file(model_answers_path)  # Optional manual editing (disabled for now)

            # Load the answers from the file after editing
            self.read_model_answers_file(model_answers_path)

            QMessageBox.information(self, "Success", "Model answers saved successfully!")

            # Disable save answers button
            self.btn_save_model_answers.setEnabled(False)

        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred: {str(e)}. Recreating the file.")
            try:
                if os.path.exists(model_answers_path):
                    os.remove(model_answers_path)
                self.create_model_answers_file(model_answers_path)
                self.open_csv_file(model_answers_path)
                QMessageBox.information(self, "Notice", "Please add the answers again.")
            except Exception as recreate_error:
                QMessageBox.critical(self, "Critical Error", f"Failed to recreate the file: {str(recreate_error)}")


    def create_model_answers_file(self, file_path):
        """Create an Excel file where only the second row is editable with allowed values 0-4"""
        wb = Workbook()
        ws = wb.active

        # Write header (Q1 to Q50) in row 1
        for col in range(1, 51):
            cell = ws.cell(row=1, column=col, value=f"Q{col}")
            cell.protection = Protection(locked=True)  # lock header

        # Write default answers (0s) in row 2
        for col in range(1, 51):
            cell = ws.cell(row=2, column=col, value=0)
            cell.protection = Protection(locked=False)  # allow editing

        # Add data validation (0 to 4) for editable answer cells
        dv = DataValidation(type="whole", operator="between", formula1=0, formula2=4)
        dv.error = "Please enter a number between 0 and 4."
        dv.errorTitle = "Invalid Input"
        ws.add_data_validation(dv)
        dv.add("A2:AX2")  # 50 columns = A to AX

        # Lock sheet
        ws.protection.sheet = True
        ws.protection.enable()

        # Save the file
        wb.save(file_path)

    def open_csv_file(self, file_path):
        """Open the XLSX file directly for editing"""
        self.btn_save_model_answers.setEnabled(True)

        if os.name == 'nt':  # Windows
            os.startfile(file_path)
        else:  # macOS or Linux
            subprocess.call(('open' if os.name == 'posix' else 'xdg-open', file_path))


    def read_model_answers_file(self, file_path):
        """Read the model answers from the XLSX file and update the processor"""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        row = [ws.cell(row=2, column=col).value for col in range(1, 51)]
        self.model_answers = [int(ans) if ans is not None else 0 for ans in row]


    def edit_answers(self):
        """Open the model answers XLSX file for editing"""
        if not self.project_path:
            QMessageBox.warning(self, "Error", "No project is open. Please select a project first.")
            return

        model_answers_path = os.path.join(self.project_path, "model_answers.xlsx")
        if os.path.exists(model_answers_path):
            self.open_csv_file(model_answers_path)
        else:
            QMessageBox.warning(self, "Error", "Model answers file does not exist. Please load it first.")
